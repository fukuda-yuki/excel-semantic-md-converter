"""Read visible workbook cell content for Phase 1 inspection."""

from __future__ import annotations

from contextlib import ExitStack
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles.numbers import is_date_format
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

from excel_semantic_md.models import SCHEMA_VERSION

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
NS = {
    "main": MAIN_NS,
    "rel": REL_NS,
    "pkg": PKG_REL_NS,
}


@dataclass
class CellReadValue:
    row: int
    col: int
    a1: str
    text: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "row": self.row,
            "col": self.col,
            "a1": self.a1,
            "text": self.text,
        }


@dataclass
class MergedRange:
    start_row: int
    start_col: int
    end_row: int
    end_col: int
    a1: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "start_row": self.start_row,
            "start_col": self.start_col,
            "end_row": self.end_row,
            "end_col": self.end_col,
            "a1": self.a1,
        }


@dataclass
class ReadWarning:
    code: str
    message: str
    details: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "code": self.code,
            "message": self.message,
            "details": dict(self.details),
        }


@dataclass
class ReadFailure:
    stage: str
    code: str
    message: str
    details: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "stage": self.stage,
            "code": self.code,
            "message": self.message,
            "details": dict(self.details),
        }


@dataclass
class SheetReadResult:
    sheet_index: int
    name: str
    cells: list[CellReadValue] = field(default_factory=list)
    merged_ranges: list[MergedRange] = field(default_factory=list)
    warnings: list[ReadWarning] = field(default_factory=list)
    failures: list[ReadFailure] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_index": self.sheet_index,
            "name": self.name,
            "cells": [cell.to_dict() for cell in self.cells],
            "merged_ranges": [merged_range.to_dict() for merged_range in self.merged_ranges],
            "warnings": [warning.to_dict() for warning in self.warnings],
            "failures": [failure.to_dict() for failure in self.failures],
        }


@dataclass
class WorkbookReadResult:
    input_file_name: str
    sheets: list[SheetReadResult] = field(default_factory=list)
    schema_version: str = SCHEMA_VERSION

    def to_dict(self) -> dict[str, Any]:
        return {
            "schema_version": self.schema_version,
            "input_file_name": self.input_file_name,
            "sheets": [sheet.to_dict() for sheet in self.sheets],
        }


@dataclass
class SheetXmlMetadata:
    sheet_index: int
    name: str
    hidden_rows: set[int] = field(default_factory=set)
    hidden_cols: set[int] = field(default_factory=set)
    merged_ranges: list[MergedRange] = field(default_factory=list)
    formula_cache_presence: dict[str, bool] = field(default_factory=dict)


def read_workbook(path: str | Path) -> WorkbookReadResult:
    """Read visible cell content without saving or modifying the input workbook."""

    workbook_path = Path(path)
    with ExitStack() as stack:
        sheet_metadata = _load_sheet_metadata(workbook_path)
        data_workbook = load_workbook(
            filename=workbook_path,
            read_only=True,
            data_only=True,
            keep_vba=False,
            keep_links=False,
        )
        stack.callback(data_workbook.close)
        result = WorkbookReadResult(input_file_name=workbook_path.name)
        for metadata in sheet_metadata:
            data_sheet = data_workbook[metadata.name]
            result.sheets.append(_read_sheet(data_sheet, metadata))
        return result


def _read_sheet(data_sheet: Any, metadata: SheetXmlMetadata) -> SheetReadResult:
    sheet = SheetReadResult(
        sheet_index=metadata.sheet_index,
        name=metadata.name,
        merged_ranges=list(metadata.merged_ranges),
    )

    formula_cache_failures = _formula_cache_failures(metadata)
    if formula_cache_failures:
        sheet.failures.extend(formula_cache_failures)
        return sheet

    for row in data_sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if cell.row in metadata.hidden_rows or cell.column in metadata.hidden_cols:
                continue
            if not _is_merged_top_left(metadata.merged_ranges, cell.coordinate):
                continue
            text = _display_text(cell)
            if text == "":
                continue
            sheet.cells.append(
                CellReadValue(
                    row=cell.row,
                    col=cell.column,
                    a1=cell.coordinate,
                    text=text,
                )
            )

    return sheet


def _formula_cache_failures(metadata: SheetXmlMetadata) -> list[ReadFailure]:
    return [
        ReadFailure(
            stage="workbook_reading",
            code="formula_cached_value_missing",
            message="Formula cell has no saved display value cache.",
            details={"cell": coordinate},
        )
        for coordinate, has_cache in metadata.formula_cache_presence.items()
        if not has_cache
    ]


def _is_merged_top_left(merged_ranges: list[MergedRange], coordinate: str) -> bool:
    for merged_range in merged_ranges:
        if not _coordinate_in_range(coordinate, merged_range.a1):
            continue
        min_col, min_row, _max_col, _max_row = range_boundaries(merged_range.a1)
        return coordinate == f"{get_column_letter(min_col)}{min_row}"
    return True


def _load_sheet_metadata(workbook_path: Path) -> list[SheetXmlMetadata]:
    import zipfile

    with zipfile.ZipFile(workbook_path) as archive:
        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        rels_root = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        relationship_targets = _relationship_targets(rels_root)
        sheet_metadata: list[SheetXmlMetadata] = []

        sheets_root = workbook_root.find("main:sheets", NS)
        if sheets_root is None:
            raise KeyError("Workbook XML has no sheets collection.")

        for sheet_index, sheet_node in enumerate(sheets_root.findall("main:sheet", NS), start=1):
            state = sheet_node.attrib.get("state", "visible")
            if state != "visible":
                continue
            rel_id = sheet_node.attrib[f"{{{REL_NS}}}id"]
            sheet_path = relationship_targets[rel_id]
            sheet_metadata.append(
                _parse_sheet_xml(
                    archive=archive,
                    sheet_index=sheet_index,
                    name=sheet_node.attrib["name"],
                    sheet_path=sheet_path,
                )
            )

        return sheet_metadata


def _relationship_targets(rels_root: ElementTree.Element) -> dict[str, str]:
    targets: dict[str, str] = {}
    for rel_node in rels_root.findall("pkg:Relationship", NS):
        rel_id = rel_node.attrib["Id"]
        target = rel_node.attrib["Target"]
        normalized = target.lstrip("/")
        if not normalized.startswith("xl/"):
            normalized = f"xl/{normalized}"
        targets[rel_id] = normalized
    return targets


def _parse_sheet_xml(*, archive: Any, sheet_index: int, name: str, sheet_path: str) -> SheetXmlMetadata:
    root = ElementTree.fromstring(archive.read(sheet_path))
    hidden_rows = _xml_hidden_rows(root)
    hidden_cols = _xml_hidden_columns(root)
    merged_ranges = _xml_merged_ranges(root, hidden_rows, hidden_cols)
    formula_cache_presence = _xml_formula_cache_presence(root, hidden_rows, hidden_cols)
    return SheetXmlMetadata(
        sheet_index=sheet_index,
        name=name,
        hidden_rows=hidden_rows,
        hidden_cols=hidden_cols,
        merged_ranges=merged_ranges,
        formula_cache_presence=formula_cache_presence,
    )


def _xml_hidden_rows(root: ElementTree.Element) -> set[int]:
    hidden: set[int] = set()
    for row_node in root.findall(".//main:sheetData/main:row", NS):
        if row_node.attrib.get("hidden") != "1":
            continue
        hidden.add(int(row_node.attrib["r"]))
    return hidden


def _xml_hidden_columns(root: ElementTree.Element) -> set[int]:
    hidden: set[int] = set()
    for col_node in root.findall(".//main:cols/main:col", NS):
        if col_node.attrib.get("hidden") != "1":
            continue
        min_col = int(col_node.attrib["min"])
        max_col = int(col_node.attrib["max"])
        hidden.update(range(min_col, max_col + 1))
    return hidden


def _xml_merged_ranges(
    root: ElementTree.Element,
    hidden_rows: set[int],
    hidden_cols: set[int],
) -> list[MergedRange]:
    merged_ranges: list[MergedRange] = []
    for merged_node in root.findall(".//main:mergeCells/main:mergeCell", NS):
        a1 = merged_node.attrib["ref"]
        min_col, min_row, max_col, max_row = range_boundaries(a1)
        if any(row in hidden_rows for row in range(min_row, max_row + 1)):
            continue
        if any(col in hidden_cols for col in range(min_col, max_col + 1)):
            continue
        merged_ranges.append(
            MergedRange(
                start_row=min_row,
                start_col=min_col,
                end_row=max_row,
                end_col=max_col,
                a1=a1,
            )
        )
    return merged_ranges


def _xml_formula_cache_presence(
    root: ElementTree.Element,
    hidden_rows: set[int],
    hidden_cols: set[int],
) -> dict[str, bool]:
    formula_cache: dict[str, bool] = {}
    for cell_node in root.findall(".//main:sheetData/main:row/main:c", NS):
        coordinate = cell_node.attrib.get("r")
        if not coordinate:
            continue
        row, col = _coordinate_parts(coordinate)
        if row in hidden_rows or col in hidden_cols:
            continue
        if cell_node.find("main:f", NS) is None:
            continue
        formula_cache[coordinate] = cell_node.find("main:v", NS) is not None
    return formula_cache


def _coordinate_parts(coordinate: str) -> tuple[int, int]:
    for index, char in enumerate(coordinate):
        if char.isdigit():
            return int(coordinate[index:]), column_index_from_string(coordinate[:index])
    raise ValueError(f"Invalid cell coordinate: {coordinate}")


def _coordinate_in_range(coordinate: str, a1_range: str) -> bool:
    row, col = _coordinate_parts(coordinate)
    min_col, min_row, max_col, max_row = range_boundaries(a1_range)
    return min_row <= row <= max_row and min_col <= col <= max_col


def _display_text(cell: Cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, str):
        return "" if value.strip() == "" else value
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return _format_datetime(value, cell.number_format)
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, (int, float, Decimal)):
        return _format_number(value, cell.number_format)
    return str(value)


def _format_datetime(value: datetime, number_format: str | None) -> str:
    if number_format and is_date_format(number_format):
        if value.time() == time(0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    return value.isoformat(sep=" ")


def _format_number(value: int | float | Decimal, number_format: str | None) -> str:
    if number_format and "%" in number_format:
        decimal_value = Decimal(str(value)) * Decimal("100")
        places = _decimal_places(number_format)
        return f"{decimal_value:.{places}f}%"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _decimal_places(number_format: str) -> int:
    percent_index = number_format.find("%")
    candidate = number_format[:percent_index] if percent_index >= 0 else number_format
    if "." not in candidate:
        return 0
    return sum(1 for char in candidate.split(".", 1)[1] if char in {"0", "#"})


__all__ = [
    "CellReadValue",
    "MergedRange",
    "ReadFailure",
    "ReadWarning",
    "SheetReadResult",
    "WorkbookReadResult",
    "read_workbook",
]

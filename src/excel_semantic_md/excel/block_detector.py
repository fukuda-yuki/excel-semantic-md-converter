"""Detect conservative Phase 1 cell-based blocks from workbook reading output."""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.utils import get_column_letter

from excel_semantic_md.excel.workbook_reader import (
    CellReadValue,
    SheetReadResult,
    WorkbookReadResult,
)
from excel_semantic_md.models import (
    HeadingBlock,
    ParagraphBlock,
    Rect,
    SheetModel,
    SourceKind,
    TableBlock,
    WarningInfo,
    WorkbookModel,
    make_block_id,
)


@dataclass(frozen=True)
class Region:
    start_row: int
    start_col: int
    end_row: int
    end_col: int

    @property
    def height(self) -> int:
        return self.end_row - self.start_row + 1

    @property
    def width(self) -> int:
        return self.end_col - self.start_col + 1


@dataclass
class TableClassification:
    header_rows: int
    header_cols: int
    warnings: list[WarningInfo]


class SheetContext:
    def __init__(self, sheet: SheetReadResult) -> None:
        self.sheet = sheet
        self.cells_by_position = {(cell.row, cell.col): cell for cell in sheet.cells}
        self.merged_by_start = {
            (merged.start_row, merged.start_col): merged
            for merged in sheet.merged_ranges
            if (merged.start_row, merged.start_col) in self.cells_by_position
        }
        self.covered_positions = self._build_covered_positions()
        self.used_positions = set(self.cells_by_position) | self.covered_positions

    def _build_covered_positions(self) -> set[tuple[int, int]]:
        covered: set[tuple[int, int]] = set()
        for merged in self.merged_by_start.values():
            for row in range(merged.start_row, merged.end_row + 1):
                for col in range(merged.start_col, merged.end_col + 1):
                    covered.add((row, col))
        return covered

    def has_used_content(self, row: int, col: int) -> bool:
        return (row, col) in self.used_positions

    def has_raw_cell(self, row: int, col: int) -> bool:
        return (row, col) in self.cells_by_position

    def cell(self, row: int, col: int) -> CellReadValue | None:
        return self.cells_by_position.get((row, col))


def detect_blocks(read_result: WorkbookReadResult) -> WorkbookModel:
    """Return the workbook block model for the conservative Phase 1 detector."""

    sheets: list[SheetModel] = []
    for read_sheet in read_result.sheets:
        context = SheetContext(read_sheet)
        blocks = _detect_sheet_blocks(context) if not read_sheet.failures else []
        sorted_blocks = sorted(
            blocks,
            key=lambda block: (
                block.anchor.start_row,
                block.anchor.start_col,
                block.anchor.end_row,
                block.anchor.end_col,
            ),
        )
        for block_index, block in enumerate(sorted_blocks, start=1):
            block.id = make_block_id(read_sheet.sheet_index, block_index, block.kind)
        sheets.append(
            SheetModel(
                sheet_index=read_sheet.sheet_index,
                name=read_sheet.name,
                blocks=sorted_blocks,
            )
        )
    return WorkbookModel(sheets=sheets, input_file_name=read_result.input_file_name)


def _detect_sheet_blocks(context: SheetContext) -> list[HeadingBlock | ParagraphBlock | TableBlock]:
    regions = _split_sheet_regions(context)
    blocks: list[HeadingBlock | ParagraphBlock | TableBlock] = []
    for region in regions:
        blocks.extend(_detect_region_blocks(context, region))
    return blocks


def _split_sheet_regions(context: SheetContext) -> list[Region]:
    if not context.used_positions:
        return []

    used_rows = sorted({row for row, _col in context.used_positions})
    row_segments = _continuous_segments(used_rows)
    regions: list[Region] = []
    for start_row, end_row in row_segments:
        used_cols = sorted(
            {
                col
                for row, col in context.used_positions
                if start_row <= row <= end_row
            }
        )
        for start_col, end_col in _continuous_segments(used_cols):
            candidate = Region(start_row=start_row, start_col=start_col, end_row=end_row, end_col=end_col)
            regions.extend(_split_region_by_internal_whitespace(context, candidate))
    return regions


def _detect_region_blocks(
    context: SheetContext,
    region: Region,
) -> list[HeadingBlock | ParagraphBlock | TableBlock]:
    trimmed = _trim_region(context, region)
    if trimmed is None:
        return []

    table_classification = _classify_table(context, trimmed)
    if table_classification is not None:
        return [_build_table_block(context, trimmed, table_classification)]

    split = _extract_leading_table_label(context, trimmed)
    if split is not None:
        _lead_region, lead_block, remainder_region = split
        blocks: list[HeadingBlock | ParagraphBlock | TableBlock] = [lead_block]
        if remainder_region is not None:
            blocks.extend(_detect_region_blocks(context, remainder_region))
        return blocks

    return [_build_paragraph_block(context, trimmed)]


def _continuous_segments(values: list[int]) -> list[tuple[int, int]]:
    if not values:
        return []

    segments: list[tuple[int, int]] = []
    start = values[0]
    end = values[0]
    for value in values[1:]:
        if value == end + 1:
            end = value
            continue
        segments.append((start, end))
        start = value
        end = value
    segments.append((start, end))
    return segments


def _trim_region(context: SheetContext, region: Region) -> Region | None:
    rows = [
        row
        for row in range(region.start_row, region.end_row + 1)
        if any(context.has_used_content(row, col) for col in range(region.start_col, region.end_col + 1))
    ]
    cols = [
        col
        for col in range(region.start_col, region.end_col + 1)
        if any(context.has_used_content(row, col) for row in range(region.start_row, region.end_row + 1))
    ]
    if not rows or not cols:
        return None
    return Region(start_row=rows[0], start_col=cols[0], end_row=rows[-1], end_col=cols[-1])


def _split_region_by_internal_whitespace(context: SheetContext, region: Region) -> list[Region]:
    trimmed = _trim_region(context, region)
    if trimmed is None:
        return []

    row_segments = _continuous_segments(
        [
            row
            for row in range(trimmed.start_row, trimmed.end_row + 1)
            if any(context.has_used_content(row, col) for col in range(trimmed.start_col, trimmed.end_col + 1))
        ]
    )
    if len(row_segments) > 1:
        regions: list[Region] = []
        for start_row, end_row in row_segments:
            regions.extend(
                _split_region_by_internal_whitespace(
                    context,
                    Region(
                        start_row=start_row,
                        start_col=trimmed.start_col,
                        end_row=end_row,
                        end_col=trimmed.end_col,
                    ),
                )
            )
        return regions

    col_segments = _continuous_segments(
        [
            col
            for col in range(trimmed.start_col, trimmed.end_col + 1)
            if any(context.has_used_content(row, col) for row in range(trimmed.start_row, trimmed.end_row + 1))
        ]
    )
    if len(col_segments) > 1:
        regions = []
        for start_col, end_col in col_segments:
            regions.extend(
                _split_region_by_internal_whitespace(
                    context,
                    Region(
                        start_row=trimmed.start_row,
                        start_col=start_col,
                        end_row=trimmed.end_row,
                        end_col=end_col,
                    ),
                )
            )
        return regions

    return [trimmed]


def _classify_table(context: SheetContext, region: Region) -> TableClassification | None:
    if region.height < 2 or region.width < 2:
        return None
    if any(not context.has_raw_cell(row, col) for row in range(region.start_row, region.end_row + 1) for col in range(region.start_col, region.end_col + 1)):
        return None

    first_row = [context.cell(region.start_row, col).text for col in range(region.start_col, region.end_col + 1)]
    first_col = [context.cell(row, region.start_col).text for row in range(region.start_row, region.end_row + 1)]
    data_rows = [
        [context.cell(row, col).text for col in range(region.start_col, region.end_col + 1)]
        for row in range(region.start_row + 1, region.end_row + 1)
    ]
    data_cols = [
        [context.cell(row, col).text for row in range(region.start_row, region.end_row + 1)]
        for col in range(region.start_col + 1, region.end_col + 1)
    ]

    first_row_numeric_count = sum(_looks_numeric(text) for text in first_row)
    first_col_numeric_count = sum(_looks_numeric(text) for text in first_col)
    other_rows_have_numeric = any(_looks_numeric(text) for row in data_rows for text in row)
    other_cols_have_numeric = any(_looks_numeric(text) for column in data_cols for text in column)

    header_row_candidate = (
        first_row_numeric_count == 0
        and (other_rows_have_numeric or region.width >= 3 or first_row != data_rows[0])
    )
    header_col_candidate = (
        first_col_numeric_count == 0
        and (other_cols_have_numeric or region.height >= 3 or first_col != data_cols[0])
    )
    if not header_row_candidate and not header_col_candidate:
        return None

    warnings: list[WarningInfo] = []
    if header_row_candidate and header_col_candidate:
        warnings.append(
            WarningInfo(
                code="ambiguous_header_detection",
                message="Both the first row and first column look like header candidates.",
                details={"anchor": _a1(region)},
            )
        )

    if header_row_candidate:
        return TableClassification(header_rows=1, header_cols=0, warnings=warnings)
    return TableClassification(header_rows=0, header_cols=1, warnings=warnings)


def _extract_leading_table_label(
    context: SheetContext,
    region: Region,
) -> tuple[Region, HeadingBlock | ParagraphBlock, Region | None] | None:
    if region.height < 2:
        return None

    top_row_cells = [
        context.cell(region.start_row, col)
        for col in range(region.start_col, region.end_col + 1)
        if context.has_raw_cell(region.start_row, col)
    ]
    if len(top_row_cells) != 1:
        return None

    remainder = _trim_region(
        context,
        Region(
            start_row=region.start_row + 1,
            start_col=region.start_col,
            end_row=region.end_row,
            end_col=region.end_col,
        ),
    )
    if remainder is None:
        return None

    table_classification = _classify_table(context, remainder)
    if table_classification is None:
        return None

    lead_cell = top_row_cells[0]
    lead_anchor_region = Region(
        start_row=lead_cell.row,
        start_col=lead_cell.col,
        end_row=lead_cell.row,
        end_col=lead_cell.col,
    )

    merged = context.merged_by_start.get((lead_cell.row, lead_cell.col))
    if merged is not None:
        lead_anchor_region = Region(
            start_row=merged.start_row,
            start_col=max(region.start_col, merged.start_col),
            end_row=min(region.end_row, merged.end_row),
            end_col=min(region.end_col, merged.end_col),
        )

    overlap = not (
        lead_anchor_region.end_col < remainder.start_col or lead_anchor_region.start_col > remainder.end_col
    )
    if not overlap:
        return None

    if merged is not None and merged.end_col > merged.start_col:
        warning = WarningInfo(
            code="table_caption_candidate",
            message="Merged text above the table is treated as a paragraph and preserved as a caption candidate.",
            details={"table_anchor": _a1(remainder)},
        )
        lead_block: HeadingBlock | ParagraphBlock = _build_paragraph_block(
            context,
            lead_anchor_region,
            warnings=[warning],
        )
    else:
        lead_block = HeadingBlock(
            id="pending",
            anchor=_rect(context.sheet.name, lead_anchor_region),
            source=SourceKind.CELLS,
            text=_region_text(context, lead_anchor_region),
            level=1,
        )

    remainder_region = Region(
        start_row=remainder.start_row,
        start_col=remainder.start_col,
        end_row=remainder.end_row,
        end_col=remainder.end_col,
    )
    return lead_anchor_region, lead_block, remainder_region


def _build_table_block(context: SheetContext, region: Region, classification: TableClassification) -> TableBlock:
    rows: list[list[str]] = []
    for row in range(region.start_row, region.end_row + 1):
        rows.append(
            [
                context.cell(row, col).text
                for col in range(region.start_col, region.end_col + 1)
            ]
        )
    return TableBlock(
        id="pending",
        anchor=_rect(context.sheet.name, region),
        source=SourceKind.CELLS,
        rows=rows,
        header_rows=classification.header_rows,
        header_cols=classification.header_cols,
        warnings=classification.warnings,
    )


def _build_paragraph_block(
    context: SheetContext,
    region: Region,
    warnings: list[WarningInfo] | None = None,
) -> ParagraphBlock:
    block_warnings = list(warnings or [])
    if _is_mixed_sparse_region(context, region):
        block_warnings.append(
            WarningInfo(
                code="mixed_sparse_region",
                message="The region is sparse or mixed, so it is preserved as a paragraph.",
                details={"anchor": _a1(region)},
            )
        )
    return ParagraphBlock(
        id="pending",
        anchor=_rect(context.sheet.name, region),
        source=SourceKind.CELLS,
        text=_region_text(context, region),
        warnings=block_warnings,
    )


def _is_mixed_sparse_region(context: SheetContext, region: Region) -> bool:
    if region.height == 1 and region.width == 1:
        return False
    used_count = sum(
        1
        for row in range(region.start_row, region.end_row + 1)
        for col in range(region.start_col, region.end_col + 1)
        if context.has_used_content(row, col)
    )
    return used_count < region.height * region.width


def _region_text(context: SheetContext, region: Region) -> str:
    lines: list[str] = []
    for row in range(region.start_row, region.end_row + 1):
        parts = [
            context.cell(row, col).text
            for col in range(region.start_col, region.end_col + 1)
            if context.has_raw_cell(row, col)
        ]
        if parts:
            lines.append(" ".join(parts))
    return "\n".join(lines)


def _rect(sheet_name: str, region: Region) -> Rect:
    return Rect(
        sheet=sheet_name,
        start_row=region.start_row,
        start_col=region.start_col,
        end_row=region.end_row,
        end_col=region.end_col,
        a1=_a1(region),
    )


def _a1(region: Region) -> str:
    start = f"{get_column_letter(region.start_col)}{region.start_row}"
    end = f"{get_column_letter(region.end_col)}{region.end_row}"
    return start if start == end else f"{start}:{end}"


def _looks_numeric(text: str) -> bool:
    candidate = text.strip().replace(",", "")
    if not candidate:
        return False
    if candidate.endswith("%"):
        candidate = candidate[:-1]
    try:
        float(candidate)
    except ValueError:
        return False
    return True


__all__ = ["detect_blocks"]

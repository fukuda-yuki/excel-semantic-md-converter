from __future__ import annotations

import contextlib
import io
import json
import os
import time
import zipfile
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import Workbook
from openpyxl.comments import Comment

import excel_semantic_md.cli.main as cli_main
from excel_semantic_md.excel import read_workbook


NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
ElementTree.register_namespace("", NS["main"])


def _run_cli(argv: list[str]) -> tuple[int | str | None, str, str]:
    stdout = io.StringIO()
    stderr = io.StringIO()
    with contextlib.redirect_stdout(stdout), contextlib.redirect_stderr(stderr):
        try:
            code = cli_main.main(argv)
        except SystemExit as exc:
            code = exc.code
    return code, stdout.getvalue(), stderr.getvalue()


def _save_workbook(workbook: Workbook, path: Path) -> Path:
    workbook.save(path)
    return path


def _set_formula_cached_value(path: Path, sheet_xml_path: str, coordinate: str, cached_value: str | None) -> None:
    temp_path = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            content = source.read(item.filename)
            if item.filename == sheet_xml_path:
                content = _patch_sheet_formula_cache(content, coordinate, cached_value)
            target.writestr(item, content)
    temp_path.replace(path)


def _patch_sheet_formula_cache(content: bytes, coordinate: str, cached_value: str | None) -> bytes:
    root = ElementTree.fromstring(content)
    cell = root.find(f".//main:c[@r='{coordinate}']", NS)
    assert cell is not None
    value_node = cell.find("main:v", NS)
    if value_node is None and cached_value is not None:
        value_node = ElementTree.SubElement(cell, f"{{{NS['main']}}}v")
    if value_node is not None and cached_value is None:
        cell.remove(value_node)
    elif value_node is not None:
        value_node.text = cached_value
    return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def test_reads_table_only_visible_cells(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Table"
    sheet["A1"] = "Name"
    sheet["B1"] = "Value"
    sheet["A2"] = "Alpha"
    sheet["B2"] = 10
    input_path = _save_workbook(workbook, tmp_path / "table.xlsx")

    data = read_workbook(input_path).to_dict()

    assert data["input_file_name"] == "table.xlsx"
    assert data["sheets"][0]["sheet_index"] == 1
    assert data["sheets"][0]["name"] == "Table"
    assert data["sheets"][0]["cells"] == [
        {"row": 1, "col": 1, "a1": "A1", "text": "Name"},
        {"row": 1, "col": 2, "a1": "B1", "text": "Value"},
        {"row": 2, "col": 1, "a1": "A2", "text": "Alpha"},
        {"row": 2, "col": 2, "a1": "B2", "text": "10"},
    ]


def test_preserves_sheet_index_and_excludes_hidden_sheets(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "Visible1"
    hidden = workbook.create_sheet("Hidden2")
    hidden.sheet_state = "hidden"
    visible = workbook.create_sheet("Visible3")
    visible["A1"] = "still third"
    input_path = _save_workbook(workbook, tmp_path / "sheets.xlsx")

    sheets = read_workbook(input_path).to_dict()["sheets"]

    assert [sheet["name"] for sheet in sheets] == ["Visible1", "Visible3"]
    assert [sheet["sheet_index"] for sheet in sheets] == [1, 3]


def test_excludes_hidden_rows_columns_and_filter_hidden_rows(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Visibility"
    sheet.append(["Name", "Visible", "Hidden"])
    sheet.append(["Visible Row", "keep", "hidden column"])
    sheet.append(["Filter Hidden Row", "drop", "hidden column"])
    sheet.append(["Hidden Row", "drop", "hidden column"])
    sheet.column_dimensions["C"].hidden = True
    sheet.row_dimensions[3].hidden = True
    sheet.row_dimensions[4].hidden = True
    sheet.auto_filter.ref = "A1:C4"
    input_path = _save_workbook(workbook, tmp_path / "visibility.xlsx")

    cells = read_workbook(input_path).to_dict()["sheets"][0]["cells"]

    assert cells == [
        {"row": 1, "col": 1, "a1": "A1", "text": "Name"},
        {"row": 1, "col": 2, "a1": "B1", "text": "Visible"},
        {"row": 2, "col": 1, "a1": "A2", "text": "Visible Row"},
        {"row": 2, "col": 2, "a1": "B2", "text": "keep"},
    ]


def test_reads_xlsm_without_modifying_input_file(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active["A1"] = "macro extension only"
    input_path = _save_workbook(workbook, tmp_path / "macro.xlsm")
    before_bytes = input_path.read_bytes()
    before_mtime = os.stat(input_path).st_mtime_ns
    time.sleep(0.01)

    data = read_workbook(input_path).to_dict()

    assert data["sheets"][0]["cells"][0]["text"] == "macro extension only"
    assert input_path.read_bytes() == before_bytes
    assert os.stat(input_path).st_mtime_ns == before_mtime


def test_uses_cached_formula_display_value_and_hides_formula_text(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = 2
    sheet["B1"] = 3
    sheet["C1"] = "=A1+B1"
    input_path = _save_workbook(workbook, tmp_path / "formula.xlsx")
    _set_formula_cached_value(input_path, "xl/worksheets/sheet1.xml", "C1", "5")

    data = read_workbook(input_path).to_dict()
    payload = json.dumps(data)

    assert {"row": 1, "col": 3, "a1": "C1", "text": "5"} in data["sheets"][0]["cells"]
    assert "=A1+B1" not in payload
    assert data["sheets"][0]["failures"] == []


def test_formula_without_cached_value_marks_sheet_failed_without_formula_text(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = 2
    sheet["B1"] = 3
    sheet["C1"] = "=A1+B1"
    input_path = _save_workbook(workbook, tmp_path / "formula-missing-cache.xlsx")
    _set_formula_cached_value(input_path, "xl/worksheets/sheet1.xml", "C1", None)

    sheet_data = read_workbook(input_path).to_dict()["sheets"][0]
    payload = json.dumps(sheet_data)

    assert sheet_data["cells"] == []
    assert sheet_data["failures"][0]["code"] == "formula_cached_value_missing"
    assert sheet_data["failures"][0]["details"] == {"cell": "C1"}
    assert "=A1+B1" not in payload


def test_formula_with_empty_string_cache_stays_non_failed_and_non_output(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = '=IF(1=1,"","x")'
    input_path = _save_workbook(workbook, tmp_path / "formula-empty-cache.xlsx")
    _set_formula_cached_value(input_path, "xl/worksheets/sheet1.xml", "A1", "")

    sheet_data = read_workbook(input_path).to_dict()["sheets"][0]

    assert sheet_data["failures"] == []
    assert sheet_data["cells"] == []


def test_normalizes_text_numbers_dates_percentages_and_merged_cells(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Formats"
    sheet["A1"] = "   "
    sheet["A2"] = "line1\nline2"
    sheet["A3"] = 42.0
    sheet["A4"] = 0.125
    sheet["A4"].number_format = "0.0%"
    sheet["A5"] = "Merged title"
    sheet.merge_cells("A5:B5")
    sheet["A6"] = "linked"
    sheet["A6"].hyperlink = "https://example.invalid"
    sheet["A7"] = "commented"
    sheet["A7"].comment = Comment("do not include", "tester")
    sheet["A8"] = datetime(2026, 4, 21, 0, 0, 0)
    sheet["A8"].number_format = "yyyy-mm-dd"
    input_path = _save_workbook(workbook, tmp_path / "formats.xlsx")

    sheet_data = read_workbook(input_path).to_dict()["sheets"][0]
    cells = sheet_data["cells"]
    payload = json.dumps(sheet_data)

    assert {"row": 1, "col": 1, "a1": "A1", "text": "   "} not in cells
    assert {"row": 2, "col": 1, "a1": "A2", "text": "line1\nline2"} in cells
    assert {"row": 3, "col": 1, "a1": "A3", "text": "42"} in cells
    assert {"row": 4, "col": 1, "a1": "A4", "text": "12.5%"} in cells
    assert {"row": 5, "col": 1, "a1": "A5", "text": "Merged title"} in cells
    assert {"row": 8, "col": 1, "a1": "A8", "text": "2026-04-21"} in cells
    assert not any(cell["a1"] == "B5" for cell in cells)
    assert sheet_data["merged_ranges"] == [
        {"start_row": 5, "start_col": 1, "end_row": 5, "end_col": 2, "a1": "A5:B5"}
    ]
    assert "https://example.invalid" not in payload
    assert "do not include" not in payload


def test_inspect_command_outputs_workbook_reading_json(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active["A1"] = "Inspect me"
    input_path = _save_workbook(workbook, tmp_path / "inspect.xlsx")

    code, stdout, stderr = _run_cli(["inspect", "--input", str(input_path)])

    data = json.loads(stdout)
    assert code == 0
    assert stderr == ""
    assert data["input_file_name"] == "inspect.xlsx"
    assert data["sheets"][0]["cells"] == [{"row": 1, "col": 1, "a1": "A1", "text": "Inspect me"}]


def test_inspect_rejects_malformed_sheet_xml_without_leaking_traceback_or_locking_file(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active["A1"] = "broken soon"
    input_path = _save_workbook(workbook, tmp_path / "malformed.xlsx")
    temp_path = input_path.with_suffix(".xlsx.tmp")

    with zipfile.ZipFile(input_path, "r") as source, zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            content = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                content = b"<worksheet"
            target.writestr(item, content)
    temp_path.replace(input_path)

    code, stdout, stderr = _run_cli(["inspect", "--input", str(input_path)])

    assert code == 2
    assert stdout == ""
    assert "failed to read input workbook" in stderr
    assert "Traceback" not in stderr
    renamed_path = input_path.with_name("renamed-malformed.xlsx")
    input_path.rename(renamed_path)
    assert renamed_path.exists()

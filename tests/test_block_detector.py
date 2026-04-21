from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from excel_semantic_md.excel import detect_blocks, read_workbook


def _save_workbook(workbook: Workbook, path: Path) -> Path:
    workbook.save(path)
    return path


def test_detects_simple_table_with_header_row(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Name"
    sheet["B1"] = "Value"
    sheet["A2"] = "Alpha"
    sheet["B2"] = 10

    model = detect_blocks(read_workbook(_save_workbook(workbook, tmp_path / "table.xlsx")))
    block = model.sheets[0].blocks[0].to_dict()

    assert block["kind"] == "table"
    assert block["header_rows"] == 1
    assert block["header_cols"] == 0
    assert block["rows"] == [["Name", "Value"], ["Alpha", "10"]]


def test_detects_vertical_table_with_header_column(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Sales"
    sheet["B1"] = 10
    sheet["A2"] = "Profit"
    sheet["B2"] = 4

    model = detect_blocks(read_workbook(_save_workbook(workbook, tmp_path / "vertical.xlsx")))
    block = model.sheets[0].blocks[0].to_dict()

    assert block["kind"] == "table"
    assert block["header_rows"] == 0
    assert block["header_cols"] == 1


def test_detects_heading_immediately_above_table(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Quarterly Summary"
    sheet["A2"] = "Name"
    sheet["B2"] = "Value"
    sheet["A3"] = "Alpha"
    sheet["B3"] = 10

    blocks = detect_blocks(read_workbook(_save_workbook(workbook, tmp_path / "heading.xlsx"))).sheets[0].blocks

    assert [block.kind.value for block in blocks] == ["heading", "table"]
    assert blocks[0].to_dict()["text"] == "Quarterly Summary"
    assert blocks[1].to_dict()["anchor"]["a1"] == "A2:B3"


def test_detects_merged_caption_as_paragraph_with_warning(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Revenue Table"
    sheet.merge_cells("A1:B1")
    sheet["A2"] = "Name"
    sheet["B2"] = "Value"
    sheet["A3"] = "Alpha"
    sheet["B3"] = 10

    blocks = detect_blocks(read_workbook(_save_workbook(workbook, tmp_path / "caption.xlsx"))).sheets[0].blocks

    assert [block.kind.value for block in blocks] == ["paragraph", "table"]
    assert blocks[0].to_dict()["anchor"]["a1"] == "A1:B1"
    assert [warning.code for warning in blocks[0].warnings] == ["table_caption_candidate"]


def test_detects_sparse_text_region_as_paragraph_with_warning(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Hello"
    sheet["B1"] = "world"
    sheet["A2"] = "from"
    sheet["B3"] = "sheet"

    blocks = detect_blocks(read_workbook(_save_workbook(workbook, tmp_path / "paragraph.xlsx"))).sheets[0].blocks

    assert len(blocks) == 1
    assert blocks[0].kind.value == "paragraph"
    assert blocks[0].to_dict()["text"] == "Hello world\nfrom\nsheet"
    assert [warning.code for warning in blocks[0].warnings] == ["mixed_sparse_region"]


def test_hidden_rows_and_columns_do_not_affect_detected_blocks(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Name"
    sheet["B1"] = "Value"
    sheet["A2"] = "Visible"
    sheet["B2"] = 10
    sheet["A3"] = "Hidden"
    sheet["B3"] = 99
    sheet["C1"] = "Ignore"
    sheet["C2"] = "Ignore"
    sheet.row_dimensions[3].hidden = True
    sheet.column_dimensions["C"].hidden = True

    blocks = detect_blocks(read_workbook(_save_workbook(workbook, tmp_path / "hidden.xlsx"))).sheets[0].blocks

    assert len(blocks) == 1
    assert blocks[0].to_dict()["anchor"]["a1"] == "A1:B2"


def test_block_ids_are_stable_in_top_to_bottom_left_to_right_order(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Overview"
    sheet["A3"] = "Name"
    sheet["B3"] = "Value"
    sheet["A4"] = "Alpha"
    sheet["B4"] = 10
    sheet["D1"] = "Notes"
    sheet["D2"] = "line2"

    blocks = detect_blocks(read_workbook(_save_workbook(workbook, tmp_path / "ids.xlsx"))).sheets[0].blocks

    assert [block.id for block in blocks] == [
        "s001-b001-paragraph",
        "s001-b002-paragraph",
        "s001-b003-table",
    ]
    assert [block.anchor.a1 for block in blocks] == ["A1", "D1:D2", "A3:B4"]
